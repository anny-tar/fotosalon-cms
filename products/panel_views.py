from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import F

from accounts.mixins import panel_required
from .models import Product, ProductCategory


@panel_required
def product_list(request):
    products = Product.objects.select_related('category').all()

    category_id = request.GET.get('category')
    in_stock = request.GET.get('in_stock')

    if category_id:
        products = products.filter(category_id=category_id)
    if in_stock == '1':
        products = products.filter(quantity__gt=0)
    elif in_stock == '0':
        products = products.filter(quantity=0)

    low_stock = products.filter(quantity__lte=F('min_quantity'))

    from django.core.paginator import Paginator
    paginator = Paginator(products, 20)
    page = request.GET.get('page', 1)
    products_page = paginator.get_page(page)

    context = {
        'products': products_page,
        'categories': ProductCategory.objects.all(),
        'current_category': category_id,
        'in_stock': in_stock,
        'low_stock_count': low_stock.count(),
    }
    return render(request, 'panel/products/list.html', context)


@panel_required
def product_create(request):
    if request.method == 'POST':
        return _save_product(request, instance=None)
    context = {
        'categories': ProductCategory.objects.all(),
        'product': None,
    }
    return render(request, 'panel/products/form.html', context)


@panel_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        return _save_product(request, instance=product)
    context = {
        'categories': ProductCategory.objects.all(),
        'product': product,
    }
    return render(request, 'panel/products/form.html', context)


def _save_product(request, instance):
    article = request.POST.get('article', '').strip()
    name = request.POST.get('name', '').strip()
    category_id = request.POST.get('category')
    description = request.POST.get('description', '').strip()
    price = request.POST.get('price', '0')
    quantity = request.POST.get('quantity', '0')
    min_quantity = request.POST.get('min_quantity', '1')
    is_active = request.POST.get('is_active') == 'on'
    order = request.POST.get('order', '0')
    image = request.FILES.get('image')

    if not article or not name:
        messages.error(request, 'Артикул и название обязательны.')
        return redirect('/panel/products/')

    # Проверка уникальности артикула
    qs = Product.objects.filter(article=article)
    if instance:
        qs = qs.exclude(pk=instance.pk)
    if qs.exists():
        messages.error(request, f'Товар с артикулом «{article}» уже существует.')
        return redirect('/panel/products/')

    if instance is None:
        instance = Product()

    instance.article = article
    instance.name = name
    instance.category_id = category_id or None
    instance.description = description
    instance.price = price
    instance.quantity = int(quantity)
    instance.min_quantity = int(min_quantity)
    instance.is_active = is_active
    instance.order = int(order)

    if image:
        # Конвертируем в WebP и создаём миниатюру
        from portfolio.image_service import convert_to_webp_and_thumbnail
        main_path, thumb_path = convert_to_webp_and_thumbnail(
            image,
            upload_to='products/images',
            thumb_upload_to='products/thumbnails',
        )
        instance.image = main_path
        instance.thumbnail = thumb_path

    instance.save()
    messages.success(request, f'Товар «{instance.name}» сохранён.')
    return redirect('/panel/products/')


@panel_required
def product_delete(request, pk):
    if request.method != 'POST':
        return redirect('/panel/products/')
    product = get_object_or_404(Product, pk=pk)
    name = product.name
    product.delete()
    messages.success(request, f'Товар «{name}» удалён.')
    return redirect('/panel/products/')


@panel_required
def product_export(request):
    """Экспорт остатков товаров в Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    products = Product.objects.select_related('category').order_by('article')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Остатки товаров'

    headers = ['Артикул', 'Название', 'Категория', 'Кол-во', 'Цена', 'Статус']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2563EB')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.article)
        ws.cell(row=row, column=2, value=product.name)
        ws.cell(row=row, column=3, value=product.category.name if product.category else '—')
        ws.cell(row=row, column=4, value=product.quantity)
        ws.cell(row=row, column=5, value=float(product.price))
        ws.cell(row=row, column=6, value='В наличии' if product.is_in_stock else 'Нет в наличии')

        # Подсвечиваем строки с низким остатком
        if product.is_low_stock:
            warn_fill = PatternFill(fill_type='solid', fgColor='FFF3CD')
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = warn_fill

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    ws.auto_filter.ref = f'A1:F{products.count() + 1}'

    today = timezone.now().strftime('%Y%m%d')
    filename = f'products_{today}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@panel_required
def category_list(request):
    categories = ProductCategory.objects.all()
    return render(request, 'panel/products/categories.html', {'categories': categories})


@panel_required
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        slug = request.POST.get('slug', '').strip()
        if not name or not slug:
            messages.error(request, 'Заполните все поля.')
            return redirect('/panel/products/categories/')
        if ProductCategory.objects.filter(slug=slug).exists():
            messages.error(request, f'Категория со slug «{slug}» уже существует.')
            return redirect('/panel/products/categories/')
        ProductCategory.objects.create(name=name, slug=slug)
        messages.success(request, f'Категория «{name}» создана.')
    return redirect('/panel/products/categories/')


@panel_required
def category_delete(request, pk):
    if request.method != 'POST':
        return redirect('/panel/products/categories/')
    category = get_object_or_404(ProductCategory, pk=pk)
    if category.product_set.exists():
        messages.error(request, 'Нельзя удалить категорию с товарами.')
        return redirect('/panel/products/categories/')
    category.delete()
    messages.success(request, 'Категория удалена.')
    return redirect('/panel/products/categories/')