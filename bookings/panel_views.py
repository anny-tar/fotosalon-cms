from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Q

from accounts.mixins import panel_required
from .models import Booking, BookingHistory, WorkingSlot
from services.models import Service


@panel_required
def booking_list(request):
    bookings = Booking.objects.select_related('service', 'slot').all()

    # Фильтрация
    status = request.GET.get('status')
    service_id = request.GET.get('service')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status:
        bookings = bookings.filter(status=status)
    if service_id:
        bookings = bookings.filter(service_id=service_id)
    if date_from:
        bookings = bookings.filter(created_at__date__gte=date_from)
    if date_to:
        bookings = bookings.filter(created_at__date__lte=date_to)

    # Пагинация вручную
    from django.core.paginator import Paginator
    paginator = Paginator(bookings, 20)
    page = request.GET.get('page', 1)
    bookings_page = paginator.get_page(page)

    context = {
        'bookings': bookings_page,
        'services': Service.objects.filter(is_active=True),
        'status_choices': Booking.STATUS_CHOICES,
        'current_status': status,
        'current_service': service_id,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'panel/bookings/list.html', context)


@panel_required
def booking_detail(request, pk):
    booking = get_object_or_404(
        Booking.objects.select_related('service', 'slot'),
        pk=pk
    )
    history = booking.history.select_related('changed_by').all()

    context = {
        'booking': booking,
        'history': history,
        'status_choices': Booking.STATUS_CHOICES,
        'notify_statuses': Booking.NOTIFY_STATUSES,
    }
    return render(request, 'panel/bookings/detail.html', context)


@panel_required
def booking_change_status(request, pk):
    if request.method != 'POST':
        return redirect('bookings_panel:detail', pk=pk)

    booking = get_object_or_404(Booking, pk=pk)
    new_status = request.POST.get('status')
    send_notify = request.POST.get('send_notify') == 'on'
    manager_note = request.POST.get('manager_note', '').strip()

    # Проверяем что статус допустимый
    valid_statuses = [s[0] for s in Booking.STATUS_CHOICES]
    if new_status not in valid_statuses:
        messages.error(request, 'Недопустимый статус.')
        return redirect('bookings_panel:detail', pk=pk)

    old_status = booking.status

    # Сохраняем примечание менеджера
    if manager_note:
        booking.manager_note = manager_note

    # Меняем статус
    booking.status = new_status

    # При отмене освобождаем слот (слот физически не удаляем,
    # просто статус заявки меняется — is_booked пересчитается сам)
    booking.save()

    # Записываем историю
    notified = False
    if new_status in Booking.NOTIFY_STATUSES and send_notify:
        notified = _send_status_notification(booking)

    BookingHistory.objects.create(
        booking=booking,
        old_status=old_status,
        new_status=new_status,
        changed_by=request.user,
        client_notified=notified,
    )

    # Журнал действий
    from accounts.models import UserActionLog
    UserActionLog.objects.create(
        user=request.user,
        action=f'Заявка #{booking.id}: статус изменён с «{old_status}» на «{new_status}»'
    )

    messages.success(request, f'Статус заявки изменён на «{booking.get_status_display()}».')
    return redirect('bookings_panel:detail', pk=pk)


def _send_status_notification(booking):
    """Отправляет email клиенту при смене статуса. Возвращает True если успешно."""
    from django.core.mail import send_mail
    from django.conf import settings
    try:
        subject = f'Обновление по вашей записи — {booking.service.name}'
        if booking.status == Booking.STATUS_CONFIRMED:
            body = (
                f'Здравствуйте, {booking.client_name}!\n\n'
                f'Ваша запись подтверждена.\n'
                f'Услуга: {booking.service.name}\n'
                f'Дата: {booking.slot.date.strftime("%d.%m.%Y")}\n'
                f'Время: {booking.slot.time_start.strftime("%H:%M")}\n\n'
                f'Ждём вас!'
            )
        else:
            body = (
                f'Здравствуйте, {booking.client_name}!\n\n'
                f'К сожалению, ваша запись на услугу «{booking.service.name}» отменена.\n'
                f'Если у вас есть вопросы — свяжитесь с нами.'
            )
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [booking.client_email])
        return True
    except Exception:
        return False


@panel_required
def booking_export(request):
    """Экспорт заявок в Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    bookings = Booking.objects.select_related('service', 'slot').all()

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        bookings = bookings.filter(created_at__date__gte=date_from)
    if date_to:
        bookings = bookings.filter(created_at__date__lte=date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Заявки'

    # Заголовки
    headers = [
        '№', 'Дата создания', 'Клиент', 'Телефон', 'Email',
        'Услуга', 'Дата записи', 'Время', 'Статус', 'Примечание'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2563EB')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Данные
    status_labels = dict(Booking.STATUS_CHOICES)
    for row, booking in enumerate(bookings, 2):
        ws.cell(row=row, column=1, value=booking.id)
        ws.cell(row=row, column=2, value=booking.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=3, value=booking.client_name)
        ws.cell(row=row, column=4, value=booking.client_phone)
        ws.cell(row=row, column=5, value=booking.client_email)
        ws.cell(row=row, column=6, value=booking.service.name)
        ws.cell(row=row, column=7, value=booking.slot.date.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=8, value=booking.slot.time_start.strftime('%H:%M'))
        ws.cell(row=row, column=9, value=status_labels.get(booking.status, booking.status))
        ws.cell(row=row, column=10, value=booking.manager_note)

    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Автофильтр
    ws.auto_filter.ref = f'A1:J{bookings.count() + 1}'

    # Отдаём файл
    today = timezone.now().strftime('%Y%m%d')
    filename = f'bookings_{today}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response